import streamlit as st
import os
import base64
import openpyxl

from openpyxl import Workbook

from google import genai
from google.genai import types

from dotenv import load_dotenv

load_dotenv()

# Initialize client globally or inside functions. Using global for caching.
API_KEY = os.getenv("GOOGLE_API_KEY") or os.getenv("API_KEY")
client = genai.Client(api_key=API_KEY) if API_KEY else None

def input_image_setup(file_obj, mime_type):
    """
    Prepares the image data for Gemini API (new SDK).
    Accepts a file object (BytesIO) and mime_type.
    Returns a list containing a types.Part object.
    """
    bytes_data = file_obj.getvalue()
    return [types.Part.from_bytes(data=bytes_data, mime_type=mime_type)]

def get_gemini_responses_high_temp(input, image, prompts):
    if not client:
         return ["Error: GOOGLE_API_KEY not found."]
    
    config = types.GenerateContentConfig(
        temperature=1.5,
        top_p=0.9,
        top_k=40,
        max_output_tokens=1024
    )
    
    all_responses = []
    for prompt in prompts:
        try:
             # image is expected to be a list containing the Part object
             contents = [input, image[0], prompt]
             response = client.models.generate_content(
                model='gemini-2.0-flash',
                contents=contents,
                config=config
             )
             
             if response.text:
                 all_responses.append(response.text)
        except Exception as e:
            print(f"Error in gemini response: {e}")
            all_responses.append(f"Error: {e}")

    #print(all_responses)
    return all_responses


def get_gemini_responses(input, image, prompts):
    if not client:
         return ["Error: GOOGLE_API_KEY not found."]

    config = types.GenerateContentConfig(
        temperature=0.5,
        top_p=0.9,
        top_k=40,
        max_output_tokens=2048
    )

    all_responses = []
    for prompt in prompts:
        try:
            contents = [input, image[0], prompt]
            response = client.models.generate_content(
                model='gemini-2.0-flash',
                contents=contents,
                config=config
            )
            if response.text:
                 all_responses.append(response.text)
        except Exception as e:
            print(f"Error in gemini response: {e}")
            all_responses.append(f"Error: {e}")

    print(all_responses)
    return all_responses

def get_gemini_dims_responses(input, image, prompts):
    if not client:
         return "Error: GOOGLE_API_KEY not found."

    config = types.GenerateContentConfig(
        temperature=1.0,
        top_p=0.9,
        top_k=40,
        max_output_tokens=5096*2
    )

    all_responses = []
    for prompt in prompts:
        try:
            contents = [input, image[0], prompt]
            response = client.models.generate_content(
                model='gemini-2.0-flash',
                contents=contents,
                config=config
            )
            if response.text:
                 all_responses.append(response.text)
        except Exception as e:
             print(f"Error in gemini response: {e}")
             all_responses.append(f"Error: {e}")

    return all_responses[0] if all_responses else ""

# def input_image_setup(uploaded_file):
#     if uploaded_file is not None:
#         bytes_data = uploaded_file.getvalue()
#         image_parts = [{"mime_type": uploaded_file.type, "data": bytes_data}]
#         return image_parts
#     else:
#         raise FileNotFoundError("No file uploaded")

def get_gemini_responses_multi_image(text_input, image_list, prompts):
    """
    Sends a request to the Gemini 1.5 Flash model with text, multiple images, 
    and a series of prompts.
    """
    if not client:
         return ["Error: GOOGLE_API_KEY not found."]

    config = types.GenerateContentConfig(
        temperature=0.5,
        top_p=0.9,
        top_k=40,
        max_output_tokens=1024
    )
    
    all_responses = []
    
    # Iterate through each prompt provided
    for prompt in prompts:
        # The key change is here: we construct the content list by combining
        # the initial text, all images from the image_list, and the current prompt.
        # The model can accept multiple images in the input list.
        content_parts = [text_input] + image_list + [prompt]
        
        try:
            # Generate content using the combined list of parts
            response = client.models.generate_content(
                model='gemini-2.0-flash',
                contents=content_parts,
                config=config
            )
            
            if response.text:
                all_responses.append(response.text)
        except Exception as e:
            print(f"Error in gemini response: {e}")
            all_responses.append(f"Error: {e}")

    return all_responses

def write_to_excel_fli_ear(results,filename,target_fields,fixed_values):
     if os.path.exists(filename):
        #print("started loading")
        #print()
        wb = openpyxl.load_workbook(filename)
        #print("finished")
        #print()
        ws = wb.worksheets[0]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        #print(headers)
        #print()
     else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Group ID", "Style Code", "Product Name", "Net Quantity",
            "Gemstone", "Base Metal", "Plating", "Type", "Color", "Model Name",
            "Occasion", "Theme", "Secondary Color", "Sales Package", "Pack of",
            "Piercing Required", "Silver Weight", "Diamond Certification", "Number of Pairs",
            "Diamond Weight", "Collection", "Ideal For", "Diamond Shape", "Diamond Clarity",
            "Diamond Color", "Number of Gemstones", "Setting", "Certification", "Brand"
        ]
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

     #print()
    # Combine target fields and fixed value fields
     all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
     target_columns = {field: headers[field] for field in all_fields if field in headers}
    # print(target_columns)

    # **Write output from row 6 onwards**
     row_idx = 1
     while ws.cell(row=row_idx, column=7).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1
    #print(row_idx)  
     for image_name, response , description in results:
        print(response)
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["Description"] = description
        field_values["Seller SKU ID"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

     wb.save(filename)


def write_to_excel_amz_xl(results,filename,target_fields,fixed_values):
    if os.path.exists(filename):
        #print("started loading")
        #print()
        wb = openpyxl.load_workbook(filename)
        #print("finished")
        #print()
        ws = wb.worksheets[0]

        # Read headers dynamically from row 1
        headers = {cell.value: cell.column for cell in ws[3] if cell.value}
        #print(headers)
        #print()
    else:
        wb = Workbook()
        ws = wb.active  

        # Define default headers
        default_headers = [
            "Image Name", "Type", "Color", "Gemstone", "Pearl Type",
            "Collection", "Occasion", "Piercing Required", "Number of Gemstones",
            "Earring Back Type", "Finish", "Design", "Metal Color", "Diamond Type",
            "Pearl Shape", "Pearl Color", "Search Keywords", "Key Features", "Trend",
            "Closure Type", "Sub Type", "Shape", "Ear Chain", "Earring Set Type",
            "Ornamentation Type", "Trend"
        ]
        
        headers = {name: idx + 1 for idx, name in enumerate(default_headers)}
        ws.append(default_headers)  # Write headers in row 1 if creating a new file

    #print()
    # Combine target fields and fixed value fields
    all_fields = set(target_fields).union(fixed_values.keys())

    # Get column indices for all relevant fields
    target_columns = {field: headers[field] for field in all_fields if field in headers}
    #print(target_columns)

    # **Write output from row 6 onwards**
    row_idx = 1
    while ws.cell(row=row_idx, column=2).value:  # Assuming column 1 is "Image Name" or always filled
        row_idx += 1
    #print(row_idx)  
    for image_name, response , description in results:
        # print(response)
        # answers = response.split("\n")  
        # answers = [ans.strip() for ans in answers]
        # answers.insert(0, os.path.splitext(image_name)[0])

        # print()
        # print(answers)  
        # print()

        # Create a dictionary mapping field names to values
        # print(target_fields)
        # field_values = {field: answers[i] if i < len(answers) else "" for i, field in enumerate(target_fields)}
        # print()
        # print(field_values)
        field_values = response

        # Write values only in the specified fields
        for key, val in fixed_values.items():
            field_values[key] = val

        field_values["product_description"] = description
        field_values["item_sku"] = os.path.splitext(image_name)[0]
        field_values["part_number"] = os.path.splitext(image_name)[0]

        for field, col_idx in target_columns.items():
            ws.cell(row=row_idx, column=col_idx, value=field_values.get(field, ""))

        row_idx += 1  # Move to the next row

    wb.save(filename)